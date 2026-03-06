"""File parsing functions for CSV and XLSX imports."""

import csv
import io
from collections.abc import Iterator
from typing import Any, BinaryIO, Union

from openpyxl import load_workbook


def parse_csv(
    file_content: Union[bytes, BinaryIO],
) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """Parse CSV file content into headers and a row generator.

    Uses streaming decode via TextIOWrapper to avoid holding the entire
    decoded text in memory at once. Tries UTF-8 first, falls back to Latin-1.

    Args:
        file_content: Raw CSV file bytes or a seekable binary file-like object.

    Returns:
        Tuple of (headers, row_generator) where rows are dicts keyed by header name.

    Raises:
        ValueError: If the CSV is empty or has no headers.
    """
    if isinstance(file_content, bytes):
        source: BinaryIO = io.BytesIO(file_content)
    else:
        source = file_content

    reader = None
    text_stream = None
    for encoding in ("utf-8", "latin-1"):
        try:
            source.seek(0)
            text_stream = io.TextIOWrapper(source, encoding=encoding)
            reader = csv.DictReader(text_stream)
            # Force header read to trigger any decode error early
            _ = reader.fieldnames
            break
        except (UnicodeDecodeError, csv.Error):
            if text_stream is not None:
                text_stream.detach()  # Don't close underlying source
            reader = None
            text_stream = None
            continue

    if reader is None or reader.fieldnames is None:
        raise ValueError("CSV file has no headers")

    headers = [h.strip() for h in reader.fieldnames if h and h.strip()]
    if not headers:
        raise ValueError("CSV file has no valid headers")

    def _row_generator() -> Iterator[dict[str, Any]]:
        for row in reader:
            cleaned = {h.strip(): str(v).strip() if v else "" for h, v in row.items() if h and h.strip()}
            if any(v for v in cleaned.values()):
                yield cleaned

    return headers, _row_generator()


def parse_xlsx(
    file_content: Union[bytes, BinaryIO],
) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """Parse XLSX file content into headers and a row generator (first sheet only).

    Uses openpyxl read_only mode and iterates rows lazily to avoid loading
    the entire sheet into memory at once. The workbook is closed when the
    generator is exhausted or garbage-collected.

    Args:
        file_content: Raw XLSX file bytes or a binary file-like object.

    Returns:
        Tuple of (headers, row_generator) where rows are dicts keyed by header name.

    Raises:
        ValueError: If the XLSX is empty or has no headers.
    """
    if isinstance(file_content, bytes):
        source: BinaryIO = io.BytesIO(file_content)
    else:
        source = file_content

    wb = load_workbook(filename=source, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("XLSX file has no worksheets")

    # Iterate rows lazily — don't call list() on the whole sheet
    row_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        wb.close()
        raise ValueError("XLSX file is empty")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    headers = [h for h in headers if h]
    if not headers:
        wb.close()
        raise ValueError("XLSX file has no valid headers")

    def _row_generator() -> Iterator[dict[str, Any]]:
        try:
            for row_values in row_iter:
                row_dict: dict[str, Any] = {}
                for j, header in enumerate(headers):
                    val = row_values[j] if j < len(row_values) else None
                    row_dict[header] = str(val).strip() if val is not None else ""
                if any(v for v in row_dict.values()):
                    yield row_dict
        finally:
            wb.close()

    return headers, _row_generator()
