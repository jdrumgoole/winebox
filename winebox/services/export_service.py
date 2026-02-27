"""Export service for generating wine and transaction exports in various formats."""

import csv
import io
from datetime import datetime
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from winebox.schemas.export import (
    ExportFormat,
    ExportMetadata,
    TransactionFlatExport,
    WineFlatExport,
)


def _format_datetime(dt: datetime | None) -> str:
    """Format datetime for export."""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _generate_filename(export_type: str, export_format: ExportFormat) -> str:
    """Generate a standardized filename for exports.

    Args:
        export_type: Type of export (e.g., "wines", "transactions")
        export_format: Export format

    Returns:
        Filename string
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    extension = export_format.value
    return f"winebox_{export_type}_{timestamp}.{extension}"


def _wine_to_row(wine: WineFlatExport) -> list[Any]:
    """Convert a wine export schema to a row for CSV/Excel."""
    return [
        wine.id,
        wine.name,
        wine.winery or "",
        wine.vintage or "",
        wine.grape_variety or "",
        wine.region or "",
        wine.country or "",
        wine.alcohol_percentage or "",
        wine.wine_type_id or "",
        wine.price_tier or "",
        wine.quantity,
        _format_datetime(wine.inventory_updated_at),
        wine.grape_blend_summary or "",
        wine.scores_summary or "",
        wine.average_score or "",
        wine.custom_fields or "",
        _format_datetime(wine.created_at),
        _format_datetime(wine.updated_at),
    ]


def _transaction_to_row(txn: TransactionFlatExport) -> list[Any]:
    """Convert a transaction export schema to a row for CSV/Excel."""
    return [
        txn.id,
        txn.wine_id,
        txn.wine_name or "",
        txn.wine_vintage or "",
        txn.wine_winery or "",
        txn.transaction_type,
        txn.quantity,
        txn.notes or "",
        _format_datetime(txn.transaction_date),
        _format_datetime(txn.created_at),
    ]


# Wine export headers
WINE_HEADERS = [
    "id",
    "name",
    "winery",
    "vintage",
    "grape_variety",
    "region",
    "country",
    "alcohol_percentage",
    "wine_type_id",
    "price_tier",
    "quantity",
    "inventory_updated_at",
    "grape_blend_summary",
    "scores_summary",
    "average_score",
    "custom_fields",
    "created_at",
    "updated_at",
]

# Transaction export headers
TRANSACTION_HEADERS = [
    "id",
    "wine_id",
    "wine_name",
    "wine_vintage",
    "wine_winery",
    "transaction_type",
    "quantity",
    "notes",
    "transaction_date",
    "created_at",
]


def export_wines_to_csv(wines: list[WineFlatExport]) -> bytes:
    """Export wines to CSV format.

    Args:
        wines: List of flat wine export schemas

    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(WINE_HEADERS)

    # Write data rows
    for wine in wines:
        writer.writerow(_wine_to_row(wine))

    return output.getvalue().encode("utf-8")


def export_wines_to_xlsx(wines: list[WineFlatExport]) -> bytes:
    """Export wines to Excel (XLSX) format.

    Args:
        wines: List of flat wine export schemas

    Returns:
        XLSX content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Wines"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B1A4A", end_color="8B1A4A", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Write header
    for col_idx, header in enumerate(WINE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, wine in enumerate(wines, 2):
        for col_idx, value in enumerate(_wine_to_row(wine), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(WINE_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        # Start with header length, then check data
        max_length = len(header)
        for row_idx in range(2, len(wines) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_wines_to_yaml(
    wines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export wines to YAML format with metadata.

    Args:
        wines: List of wine dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "wines": wines,
        "export_info": {
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "total_count": len(wines),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_wines_to_json(
    wines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export wines to JSON format with metadata.

    Args:
        wines: List of wine dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "wines": wines,
        "export_info": ExportMetadata(
            total_count=len(wines),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }


def export_transactions_to_csv(transactions: list[TransactionFlatExport]) -> bytes:
    """Export transactions to CSV format.

    Args:
        transactions: List of flat transaction export schemas

    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(TRANSACTION_HEADERS)

    # Write data rows
    for txn in transactions:
        writer.writerow(_transaction_to_row(txn))

    return output.getvalue().encode("utf-8")


def export_transactions_to_xlsx(transactions: list[TransactionFlatExport]) -> bytes:
    """Export transactions to Excel (XLSX) format.

    Args:
        transactions: List of flat transaction export schemas

    Returns:
        XLSX content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B1A4A", end_color="8B1A4A", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Write header
    for col_idx, header in enumerate(TRANSACTION_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, txn in enumerate(transactions, 2):
        for col_idx, value in enumerate(_transaction_to_row(txn), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(TRANSACTION_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(transactions) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_transactions_to_yaml(
    transactions: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export transactions to YAML format with metadata.

    Args:
        transactions: List of transaction dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "transactions": transactions,
        "export_info": {
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "total_count": len(transactions),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_transactions_to_json(
    transactions: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export transactions to JSON format with metadata.

    Args:
        transactions: List of transaction dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "transactions": transactions,
        "export_info": ExportMetadata(
            total_count=len(transactions),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }


def get_content_type(export_format: ExportFormat) -> str:
    """Get the MIME type for an export format.

    Args:
        export_format: Export format

    Returns:
        MIME type string
    """
    content_types = {
        ExportFormat.CSV: "text/csv",
        ExportFormat.XLSX: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ExportFormat.YAML: "application/x-yaml",
        ExportFormat.JSON: "application/json",
    }
    return content_types[export_format]


# X-Wines export headers
XWINES_HEADERS = [
    "id",
    "name",
    "winery",
    "wine_type",
    "country",
    "region",
    "abv",
    "avg_rating",
    "rating_count",
]


def _xwine_to_row(xwine: dict[str, Any]) -> list[Any]:
    """Convert an X-Wines result dictionary to a row for CSV/Excel."""
    return [
        xwine.get("id", ""),
        xwine.get("name", ""),
        xwine.get("winery", "") or "",
        xwine.get("wine_type", "") or "",
        xwine.get("country", "") or "",
        xwine.get("region", "") or "",
        xwine.get("abv", "") or "",
        xwine.get("avg_rating", "") or "",
        xwine.get("rating_count", 0),
    ]


def export_xwines_to_csv(xwines: list[dict[str, Any]]) -> bytes:
    """Export X-Wines search results to CSV format.

    Args:
        xwines: List of X-Wines result dictionaries

    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(XWINES_HEADERS)

    # Write data rows
    for xwine in xwines:
        writer.writerow(_xwine_to_row(xwine))

    return output.getvalue().encode("utf-8")


def export_xwines_to_xlsx(xwines: list[dict[str, Any]]) -> bytes:
    """Export X-Wines search results to Excel (XLSX) format.

    Args:
        xwines: List of X-Wines result dictionaries

    Returns:
        XLSX content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "X-Wines"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B1A4A", end_color="8B1A4A", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Write header
    for col_idx, header in enumerate(XWINES_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, xwine in enumerate(xwines, 2):
        for col_idx, value in enumerate(_xwine_to_row(xwine), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(XWINES_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(xwines) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_xwines_to_yaml(
    xwines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export X-Wines search results to YAML format with metadata.

    Args:
        xwines: List of X-Wines result dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "xwines": xwines,
        "export_info": {
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "total_count": len(xwines),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_xwines_to_json(
    xwines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export X-Wines search results to JSON format with metadata.

    Args:
        xwines: List of X-Wines result dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "xwines": xwines,
        "export_info": ExportMetadata(
            total_count=len(xwines),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }


def generate_xwines_filename(export_format: ExportFormat) -> str:
    """Generate a standardized filename for X-Wines exports.

    Args:
        export_format: Export format

    Returns:
        Filename string
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    extension = export_format.value
    return f"xwines_search_{timestamp}.{extension}"
