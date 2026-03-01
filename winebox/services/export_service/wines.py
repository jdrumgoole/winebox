"""Wine export functions."""

import csv
import io
from datetime import datetime, timezone
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from winebox.schemas.export import ExportMetadata, WineFlatExport

from .constants import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    WINE_HEADERS,
    format_datetime,
)


def _wine_to_base_row(wine: WineFlatExport) -> list[Any]:
    """Convert a wine export schema to a base row (without custom fields)."""
    return [
        wine.id,
        wine.name,
        wine.winery or "",
        wine.vintage or "",
        wine.grape_variety or "",
        wine.region or "",
        wine.country or "",
        wine.alcohol_percentage or "",
        wine.wine_type_id or "",
        wine.price_tier or "",
        wine.quantity,
        format_datetime(wine.inventory_updated_at),
        wine.grape_blend_summary or "",
        wine.scores_summary or "",
        wine.average_score or "",
        format_datetime(wine.created_at),
        format_datetime(wine.updated_at),
    ]


def _collect_custom_field_keys(wines: list[WineFlatExport]) -> list[str]:
    """Collect all unique custom field keys across wines, sorted alphabetically."""
    keys: set[str] = set()
    for wine in wines:
        if wine.custom_fields_dict:
            keys.update(wine.custom_fields_dict.keys())
    return sorted(keys)


def _build_headers_and_rows(
    wines: list[WineFlatExport],
) -> tuple[list[str], list[list[Any]]]:
    """Build headers and rows with custom fields expanded into individual columns.

    Returns:
        Tuple of (headers, rows) where custom field keys are appended after base headers.
    """
    custom_keys = _collect_custom_field_keys(wines)
    headers = list(WINE_HEADERS) + custom_keys

    rows: list[list[Any]] = []
    for wine in wines:
        base_row = _wine_to_base_row(wine)
        custom_values = [
            (wine.custom_fields_dict or {}).get(key, "") for key in custom_keys
        ]
        rows.append(base_row + custom_values)

    return headers, rows


def export_wines_to_csv(wines: list[WineFlatExport]) -> bytes:
    """Export wines to CSV format.

    Args:
        wines: List of flat wine export schemas

    Returns:
        CSV content as bytes
    """
    headers, rows = _build_headers_and_rows(wines)

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(headers)

    # Write data rows
    for row in rows:
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def export_wines_to_xlsx(wines: list[WineFlatExport]) -> bytes:
    """Export wines to Excel (XLSX) format.

    Args:
        wines: List of flat wine export schemas

    Returns:
        XLSX content as bytes
    """
    headers, rows = _build_headers_and_rows(wines)

    wb = Workbook()
    ws = wb.active
    ws.title = "Wines"

    # Write header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        # Start with header length, then check data
        max_length = len(header)
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_wines_to_yaml(
    wines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export wines to YAML format with metadata.

    Args:
        wines: List of wine dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "wines": wines,
        "export_info": {
            "exported_at": datetime.now(timezone.utc).isoformat() + "Z",
            "total_count": len(wines),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_wines_to_json(
    wines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export wines to JSON format with metadata.

    Args:
        wines: List of wine dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "wines": wines,
        "export_info": ExportMetadata(
            total_count=len(wines),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }
