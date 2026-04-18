"""X-Wines export functions."""

import csv
import io
from datetime import datetime, timezone
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from winebox.schemas.export import ExportFormat, ExportMetadata

from .constants import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    XWINES_HEADERS,
    format_exported_at,
)


def _xwine_to_row(xwine: dict[str, Any]) -> list[Any]:
    """Convert an X-Wines result dictionary to a row for CSV/Excel."""
    return [
        xwine.get("id", ""),
        xwine.get("name", ""),
        xwine.get("winery", "") or "",
        xwine.get("wine_type", "") or "",
        xwine.get("country", "") or "",
        xwine.get("region", "") or "",
        xwine.get("abv", "") or "",
        xwine.get("avg_rating", "") or "",
        xwine.get("rating_count", 0),
    ]


def export_xwines_to_csv(xwines: list[dict[str, Any]]) -> bytes:
    """Export X-Wines search results to CSV format.

    Args:
        xwines: List of X-Wines result dictionaries

    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(XWINES_HEADERS)

    # Write data rows
    for xwine in xwines:
        writer.writerow(_xwine_to_row(xwine))

    return output.getvalue().encode("utf-8")


def export_xwines_to_xlsx(xwines: list[dict[str, Any]]) -> bytes:
    """Export X-Wines search results to Excel (XLSX) format.

    Args:
        xwines: List of X-Wines result dictionaries

    Returns:
        XLSX content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "X-Wines"

    # Write header
    for col_idx, header in enumerate(XWINES_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data rows
    for row_idx, xwine in enumerate(xwines, 2):
        for col_idx, value in enumerate(_xwine_to_row(xwine), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(XWINES_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(xwines) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_xwines_to_yaml(
    xwines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export X-Wines search results to YAML format with metadata.

    Args:
        xwines: List of X-Wines result dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "xwines": xwines,
        "export_info": {
            "exported_at": format_exported_at(),
            "total_count": len(xwines),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_xwines_to_json(
    xwines: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export X-Wines search results to JSON format with metadata.

    Args:
        xwines: List of X-Wines result dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "xwines": xwines,
        "export_info": ExportMetadata(
            total_count=len(xwines),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }


def generate_xwines_filename(export_format: ExportFormat) -> str:
    """Generate a standardized filename for X-Wines exports.

    Args:
        export_format: Export format

    Returns:
        Filename string
    """
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    extension = export_format.value
    return f"xwines_search_{timestamp}.{extension}"
