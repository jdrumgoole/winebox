"""Transaction export functions."""

import csv
import io
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from winebox.schemas.export import ExportMetadata, TransactionFlatExport

from .constants import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    TRANSACTION_HEADERS,
    format_datetime,
    format_exported_at,
)


def _transaction_to_row(txn: TransactionFlatExport) -> list[Any]:
    """Convert a transaction export schema to a row for CSV/Excel."""
    return [
        txn.id,
        txn.wine_id,
        txn.wine_name or "",
        txn.wine_vintage or "",
        txn.wine_winery or "",
        txn.transaction_type,
        txn.quantity,
        txn.notes or "",
        format_datetime(txn.transaction_date),
        format_datetime(txn.created_at),
        # Phase 5 — case snapshots (empty on bottle/loose/legacy rows).
        txn.item_type or "",
        txn.case_size_at_event if txn.case_size_at_event is not None else "",
        txn.provenance_at_event or "",
    ]


def export_transactions_to_csv(transactions: list[TransactionFlatExport]) -> bytes:
    """Export transactions to CSV format.

    Args:
        transactions: List of flat transaction export schemas

    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(TRANSACTION_HEADERS)

    # Write data rows
    for txn in transactions:
        writer.writerow(_transaction_to_row(txn))

    return output.getvalue().encode("utf-8")


def export_transactions_to_xlsx(transactions: list[TransactionFlatExport]) -> bytes:
    """Export transactions to Excel (XLSX) format.

    Args:
        transactions: List of flat transaction export schemas

    Returns:
        XLSX content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Write header
    for col_idx, header in enumerate(TRANSACTION_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data rows
    for row_idx, txn in enumerate(transactions, 2):
        for col_idx, value in enumerate(_transaction_to_row(txn), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(TRANSACTION_HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(transactions) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_transactions_to_yaml(
    transactions: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Export transactions to YAML format with metadata.

    Args:
        transactions: List of transaction dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        YAML content as bytes
    """
    export_data = {
        "transactions": transactions,
        "export_info": {
            "exported_at": format_exported_at(),
            "total_count": len(transactions),
            "format": "yaml",
            "filters_applied": filters_applied,
        },
    }
    return yaml.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False).encode("utf-8")


def export_transactions_to_json(
    transactions: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> dict[str, Any]:
    """Export transactions to JSON format with metadata.

    Args:
        transactions: List of transaction dictionaries
        filters_applied: Filters that were applied to the export

    Returns:
        JSON-serializable dictionary
    """
    return {
        "transactions": transactions,
        "export_info": ExportMetadata(
            total_count=len(transactions),
            format="json",
            filters_applied=filters_applied,
        ).model_dump(mode="json"),
    }
