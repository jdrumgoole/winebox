"""Import service for parsing spreadsheets and creating wine records."""

import csv
import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Any

from beanie import PydanticObjectId
from openpyxl import load_workbook

from winebox.config import settings
from winebox.models.import_batch import ImportBatch, ImportStatus
from winebox.models.wine import InventoryInfo, Wine

logger = logging.getLogger(__name__)

# Maximum rows per import batch (safety limit for MongoDB 16MB doc size)
MAX_ROWS = 5000

# Header alias table: lowercase alias -> wine field name
HEADER_ALIASES: dict[str, str] = {
    # name
    "wine": "name",
    "wine name": "name",
    "wine_name": "name",
    "label": "name",
    "name": "name",
    "title": "name",
    # winery
    "winery": "winery",
    "producer": "winery",
    "maker": "winery",
    "domaine": "winery",
    "chateau": "winery",
    "château": "winery",
    "estate": "winery",
    "bodega": "winery",
    "cantina": "winery",
    # vintage
    "vintage": "vintage",
    "year": "vintage",
    "vintage year": "vintage",
    "vintage_year": "vintage",
    # grape_variety
    "grape": "grape_variety",
    "grape variety": "grape_variety",
    "grape_variety": "grape_variety",
    "varietal": "grape_variety",
    "grapes": "grape_variety",
    "variety": "grape_variety",
    "cépage": "grape_variety",
    # region
    "region": "region",
    "wine region": "region",
    # sub_region
    "sub region": "sub_region",
    "sub_region": "sub_region",
    "subregion": "sub_region",
    # appellation
    "appellation": "appellation",
    "aoc": "appellation",
    "doc": "appellation",
    "docg": "appellation",
    "ava": "appellation",
    # country
    "country": "country",
    "origin": "country",
    "country of origin": "country",
    "country_of_origin": "country",
    # alcohol_percentage
    "alcohol": "alcohol_percentage",
    "alcohol %": "alcohol_percentage",
    "alcohol_percentage": "alcohol_percentage",
    "abv": "alcohol_percentage",
    "alcohol percentage": "alcohol_percentage",
    "alc": "alcohol_percentage",
    # wine_type_id
    "type": "wine_type_id",
    "wine type": "wine_type_id",
    "wine_type": "wine_type_id",
    "color": "wine_type_id",
    "colour": "wine_type_id",
    # classification
    "classification": "classification",
    "class": "classification",
    "grade": "classification",
    # price_tier
    "price": "price_tier",
    "price tier": "price_tier",
    "price_tier": "price_tier",
    # quantity
    "quantity": "quantity",
    "qty": "quantity",
    "bottles": "quantity",
    "count": "quantity",
    # notes (for transactions)
    "notes": "notes",
    "note": "notes",
    "description": "notes",
    "tasting notes": "notes",
}

# Valid wine field names that can be mapped to
VALID_WINE_FIELDS = {
    "name", "winery", "vintage", "grape_variety", "region", "sub_region",
    "appellation", "country", "alcohol_percentage", "wine_type_id",
    "classification", "price_tier", "quantity", "notes",
}

# Human-readable descriptions for each wine field (used in AI mapping prompt)
WINE_FIELD_DESCRIPTIONS: dict[str, str] = {
    "name": "The wine name or label title",
    "winery": "The winery, producer, domaine, château, or estate",
    "vintage": "The vintage year (integer)",
    "grape_variety": "The grape variety or varietal (e.g. Cabernet Sauvignon, Merlot)",
    "region": "The wine region (e.g. Bordeaux, Napa Valley, Rioja)",
    "sub_region": "A sub-region within the main region",
    "appellation": "The specific appellation, AOC, DOC, DOCG, or AVA",
    "country": "The country of origin",
    "alcohol_percentage": "The alcohol percentage (numeric, e.g. 13.5)",
    "wine_type_id": "The wine type: red, white, rosé, sparkling, fortified, or dessert",
    "classification": "Quality classification (e.g. Grand Cru, Reserva, DOCG)",
    "price_tier": "Price or price tier",
    "quantity": "Number of bottles (integer)",
    "notes": "Tasting notes, description, or general notes",
}

# Non-wine keywords for filtering
NON_WINE_KEYWORDS = {
    "whiskey", "whisky", "bourbon", "scotch", "cognac", "brandy",
    "gin", "vodka", "rum", "tequila", "mezcal", "beer", "ale",
    "lager", "stout", "sake", "liqueur", "liquor", "spirit",
    "spirits", "absinthe", "grappa", "armagnac", "cider",
}


def parse_csv(file_content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse CSV file content into headers and rows.

    Uses streaming decode via TextIOWrapper to avoid holding the entire
    decoded text in memory at once. Tries UTF-8 first, falls back to Latin-1.

    Args:
        file_content: Raw CSV file bytes.

    Returns:
        Tuple of (headers, rows) where rows are dicts keyed by header name.

    Raises:
        ValueError: If the CSV is empty or has no headers.
    """
    # Try each encoding with streaming TextIOWrapper (avoids decoding entire
    # file into a string at once — memory stays ~1x file size, not 3x).
    reader = None
    for encoding in ("utf-8", "latin-1"):
        try:
            text_stream = io.TextIOWrapper(io.BytesIO(file_content), encoding=encoding)
            reader = csv.DictReader(text_stream)
            # Force header read to trigger any decode error early
            _ = reader.fieldnames
            break
        except (UnicodeDecodeError, csv.Error):
            reader = None
            continue

    if reader is None or reader.fieldnames is None:
        raise ValueError("CSV file has no headers")

    headers = [h.strip() for h in reader.fieldnames if h and h.strip()]
    if not headers:
        raise ValueError("CSV file has no valid headers")

    rows: list[dict[str, Any]] = []
    for i, row in enumerate(reader):
        if i >= MAX_ROWS:
            break
        # Only include non-empty rows — process one row at a time
        cleaned = {h.strip(): str(v).strip() if v else "" for h, v in row.items() if h and h.strip()}
        if any(v for v in cleaned.values()):
            rows.append(cleaned)

    return headers, rows


def parse_xlsx(file_content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse XLSX file content into headers and rows (first sheet only).

    Uses openpyxl read_only mode and iterates rows lazily to avoid loading
    the entire sheet into memory at once.

    Args:
        file_content: Raw XLSX file bytes.

    Returns:
        Tuple of (headers, rows) where rows are dicts keyed by header name.

    Raises:
        ValueError: If the XLSX is empty or has no headers.
    """
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("XLSX file has no worksheets")

    # Iterate rows lazily — don't call list() on the whole sheet
    row_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        wb.close()
        raise ValueError("XLSX file is empty")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    headers = [h for h in headers if h]
    if not headers:
        wb.close()
        raise ValueError("XLSX file has no valid headers")

    rows: list[dict[str, Any]] = []
    for i, row_values in enumerate(row_iter):
        if i >= MAX_ROWS:
            break
        row_dict: dict[str, Any] = {}
        for j, header in enumerate(headers):
            val = row_values[j] if j < len(row_values) else None
            row_dict[header] = str(val).strip() if val is not None else ""
        if any(v for v in row_dict.values()):
            rows.append(row_dict)

    wb.close()
    return headers, rows


def suggest_column_mapping(headers: list[str]) -> dict[str, str]:
    """Auto-suggest column mapping based on header names.

    Args:
        headers: List of column header names from the spreadsheet.

    Returns:
        Dict mapping header name -> wine field name or "custom:<header>".
    """
    mapping: dict[str, str] = {}
    for header in headers:
        normalized = header.lower().strip()
        if normalized in HEADER_ALIASES:
            mapping[header] = HEADER_ALIASES[normalized]
        else:
            mapping[header] = f"custom:{header}"
    return mapping


def _static_fallback(header: str) -> str:
    """Look up a single header in the static alias table.

    Args:
        header: Column header name.

    Returns:
        Matched wine field name or "custom:<header>".
    """
    normalized = header.lower().strip()
    if normalized in HEADER_ALIASES:
        return HEADER_ALIASES[normalized]
    return f"custom:{header}"


def _build_mapping_prompt(
    headers: list[str],
    preview_rows: list[dict[str, Any]],
) -> str:
    """Build the Claude prompt for AI-assisted column mapping.

    Args:
        headers: Column header names from the spreadsheet.
        preview_rows: Up to 5 sample rows for context.

    Returns:
        Prompt string for the Claude API.
    """
    # Build field list with descriptions
    fields_section = "\n".join(
        f'  - "{field}": {desc}' for field, desc in WINE_FIELD_DESCRIPTIONS.items()
    )

    # Build header + sample values section
    header_sections: list[str] = []
    for header in headers:
        samples = []
        for row in preview_rows[:3]:
            val = row.get(header, "")
            if val:
                samples.append(str(val))
        sample_text = ", ".join(f'"{s}"' for s in samples) if samples else "(no values)"
        header_sections.append(f'  - "{header}": sample values: {sample_text}')
    headers_section = "\n".join(header_sections)

    return f"""You are mapping spreadsheet column headers to wine database fields.

Valid wine fields:
{fields_section}

Special values:
  - "skip": Ignore this column entirely
  - "custom:<name>": Store as a custom field with the given name

Spreadsheet columns (with sample values):
{headers_section}

Instructions:
- Map each column header to the most appropriate wine field, "skip", or "custom:<name>".
- Consider typos, abbreviations, and non-English headers (French, Italian, Spanish, German, etc.).
- Use sample values to disambiguate ambiguous headers (e.g. "Type" with values "Red", "White" -> "wine_type_id").
- If a column clearly doesn't match any wine field, use "custom:<original header name>".
- Return ONLY a JSON object mapping each header to its target field. No extra text.

Example output:
{{"Wine Name": "name", "Producer": "winery", "Year": "vintage", "Rating": "custom:Rating"}}"""


def _is_valid_mapping_value(value: str) -> bool:
    """Check if a mapping value is valid (known field, "skip", or "custom:...").

    Args:
        value: The mapping target value.

    Returns:
        True if the value is valid.
    """
    if value == "skip":
        return True
    if value.startswith("custom:") and len(value) > 7:
        return True
    return value in VALID_WINE_FIELDS


async def suggest_column_mapping_ai(
    headers: list[str],
    preview_rows: list[dict[str, Any]],
) -> dict[str, str] | None:
    """Suggest column mapping using Claude Haiku for smarter matching.

    Falls back to None if the API key is missing, the call fails, or the
    response can't be parsed. The caller should then use the static
    suggest_column_mapping() instead.

    Args:
        headers: Column header names from the spreadsheet.
        preview_rows: Sample rows (up to 5) for context.

    Returns:
        Dict mapping header -> wine field, or None on failure.
    """
    # Check for API key
    api_key = settings.anthropic_api_key or os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        logger.debug("No Anthropic API key available, skipping AI mapping")
        return None

    try:
        import anthropic

        client = anthropic.Anthropic(api_key=api_key)
        prompt = _build_mapping_prompt(headers, preview_rows)

        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )

        response_text = message.content[0].text

        # Handle markdown code blocks (same pattern as vision.py)
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0]
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0]

        result = json.loads(response_text.strip())

        if not isinstance(result, dict):
            logger.warning("AI mapping returned non-dict: %s", type(result).__name__)
            return None

        # Validate each mapping; fall back to static per-header for invalid ones
        validated: dict[str, str] = {}
        for header in headers:
            ai_value = result.get(header)
            if ai_value and isinstance(ai_value, str) and _is_valid_mapping_value(ai_value):
                validated[header] = ai_value
            else:
                validated[header] = _static_fallback(header)
                if ai_value is not None:
                    logger.debug(
                        "AI mapping for '%s' -> '%s' invalid, using static fallback",
                        header,
                        ai_value,
                    )

        return validated

    except json.JSONDecodeError as e:
        logger.warning("Failed to parse AI mapping response as JSON: %s", e)
        return None
    except Exception as e:
        logger.warning("AI column mapping failed: %s", e)
        return None


def is_non_wine_row(row: dict[str, Any], mapping: dict[str, str]) -> bool:
    """Check if a row appears to be a non-wine item (spirits, beer, etc.).

    Checks columns mapped to wine_type_id or name for non-wine keywords.

    Args:
        row: Raw row dict from spreadsheet.
        mapping: Column mapping dict.

    Returns:
        True if the row appears to be a non-wine item.
    """
    # Find columns mapped to type or name
    cols_to_check = []
    for header, field in mapping.items():
        if field in ("wine_type_id", "name"):
            cols_to_check.append(header)

    for col in cols_to_check:
        value = row.get(col, "").lower().strip()
        for keyword in NON_WINE_KEYWORDS:
            if keyword in value:
                return True

    return False


def _coerce_vintage(value: str) -> int | None:
    """Try to coerce a string to a vintage year."""
    if not value:
        return None
    try:
        year = int(float(value))
        if 1900 <= year <= 2100:
            return year
    except (ValueError, OverflowError):
        pass
    return None


def _coerce_float(value: str) -> float | None:
    """Try to coerce a string to a float."""
    if not value:
        return None
    # Strip % sign if present
    cleaned = value.replace("%", "").strip()
    try:
        return float(cleaned)
    except (ValueError, OverflowError):
        return None


def _coerce_int(value: str) -> int | None:
    """Try to coerce a string to an int."""
    if not value:
        return None
    try:
        return int(float(value))
    except (ValueError, OverflowError):
        return None


def _compute_custom_fields_text(custom_fields: dict[str, str] | None) -> str | None:
    """Compute denormalized text from custom fields for search indexing."""
    if not custom_fields:
        return None
    return " ".join(f"{k} {v}" for k, v in custom_fields.items())


def row_to_wine_data(
    row: dict[str, Any],
    mapping: dict[str, str],
    owner_id: PydanticObjectId,
    default_quantity: int = 1,
) -> dict[str, Any] | None:
    """Convert a spreadsheet row to Wine constructor kwargs.

    Args:
        row: Raw row dict from spreadsheet.
        mapping: Column mapping dict.
        owner_id: Owner's ID.
        default_quantity: Default quantity if not specified in row.

    Returns:
        Dict of Wine constructor kwargs, or None if row has no name.
    """
    wine_data: dict[str, Any] = {}
    custom_fields: dict[str, str] = {}
    quantity = default_quantity

    for header, field in mapping.items():
        value = row.get(header, "").strip()
        if not value:
            continue

        if field == "skip":
            continue
        elif field.startswith("custom:"):
            custom_field_name = field[7:]  # Remove "custom:" prefix
            custom_fields[custom_field_name] = value
        elif field == "vintage":
            coerced = _coerce_vintage(value)
            if coerced is not None:
                wine_data["vintage"] = coerced
        elif field == "alcohol_percentage":
            coerced = _coerce_float(value)
            if coerced is not None:
                wine_data["alcohol_percentage"] = coerced
        elif field == "quantity":
            coerced = _coerce_int(value)
            if coerced is not None and coerced > 0:
                quantity = coerced
        elif field in VALID_WINE_FIELDS:
            wine_data[field] = value

    # Name is required
    if "name" not in wine_data:
        return None

    wine_data["owner_id"] = owner_id
    wine_data["front_label_text"] = ""
    wine_data["inventory"] = InventoryInfo(
        quantity=quantity,
        updated_at=datetime.now(timezone.utc),
    )

    if custom_fields:
        wine_data["custom_fields"] = custom_fields
        wine_data["custom_fields_text"] = _compute_custom_fields_text(custom_fields)

    return wine_data


async def process_import_batch(
    batch: ImportBatch,
    owner_id: PydanticObjectId,
    skip_non_wine: bool = True,
    default_quantity: int = 1,
) -> ImportBatch:
    """Process an import batch: create Wine documents from mapped rows.

    Args:
        batch: The ImportBatch document with rows and mapping.
        owner_id: Owner's ID for created wines.
        skip_non_wine: Whether to skip non-wine rows.
        default_quantity: Default bottle quantity.

    Returns:
        Updated ImportBatch with processing results.
    """
    if not batch.column_mapping:
        batch.status = ImportStatus.FAILED
        batch.errors.append("No column mapping set")
        await batch.save()
        return batch

    batch.status = ImportStatus.PROCESSING
    await batch.save()

    wines_created = 0
    rows_skipped = 0
    errors: list[str] = []

    for i, row in enumerate(batch.rows):
        try:
            # Skip non-wine rows
            if skip_non_wine and is_non_wine_row(row, batch.column_mapping):
                rows_skipped += 1
                continue

            wine_data = row_to_wine_data(row, batch.column_mapping, owner_id, default_quantity)
            if wine_data is None:
                rows_skipped += 1
                continue

            wine = Wine(**wine_data)
            await wine.insert()
            wines_created += 1

        except Exception as e:
            error_msg = f"Row {i + 1}: {str(e)}"
            errors.append(error_msg)
            logger.warning("Import error on row %d: %s", i + 1, e)

    batch.wines_created = wines_created
    batch.rows_skipped = rows_skipped
    batch.errors = errors
    batch.status = ImportStatus.COMPLETED
    await batch.save()

    return batch
