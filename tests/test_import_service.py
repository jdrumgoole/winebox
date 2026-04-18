"""Unit tests for import service parsing, mapping, and filtering."""

import csv
import io
import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from winebox.services.import_service import (
    CANONICAL_WINE_FIELDS,
    VALID_WINE_FIELDS,
    _coerce_float,
    _coerce_vintage,
    _compute_custom_fields_text,
    _is_valid_mapping_value,
    _static_fallback,
    is_non_wine_row,
    parse_csv,
    parse_xlsx,
    row_to_wine_data,
    suggest_column_mapping,
    suggest_column_mapping_ai,
)


# =============================================================================
# CSV Parsing Tests
# =============================================================================


def test_parse_csv_basic() -> None:
    """Test basic CSV parsing."""
    content = "Wine Name,Vintage,Country\nChateau Margaux,2015,France\nBarolo,2018,Italy\n"
    headers, row_gen = parse_csv(content.encode("utf-8"))
    rows = list(row_gen)
    assert headers == ["Wine Name", "Vintage", "Country"]
    assert len(rows) == 2
    assert rows[0]["Wine Name"] == "Chateau Margaux"
    assert rows[1]["Country"] == "Italy"


def test_parse_csv_empty_rows() -> None:
    """Test that empty rows are skipped."""
    content = "Name,Vintage\nWine A,2020\n,,\nWine B,2019\n"
    headers, row_gen = parse_csv(content.encode("utf-8"))
    rows = list(row_gen)
    assert len(rows) == 2
    assert rows[0]["Name"] == "Wine A"
    assert rows[1]["Name"] == "Wine B"


def test_parse_csv_latin1_encoding() -> None:
    """Test CSV with Latin-1 encoding (accented characters)."""
    content = "Name,Region\nChâteau Lafite,Médoc\n"
    headers, row_gen = parse_csv(content.encode("latin-1"))
    rows = list(row_gen)
    assert len(rows) == 1
    assert rows[0]["Name"] == "Château Lafite"
    assert rows[0]["Region"] == "Médoc"


def test_parse_csv_no_headers() -> None:
    """Test error on CSV with no headers."""
    with pytest.raises(ValueError, match="no headers"):
        parse_csv(b"")


def test_parse_csv_headers_only() -> None:
    """Test CSV with headers but no data rows."""
    content = "Name,Vintage\n"
    headers, row_gen = parse_csv(content.encode("utf-8"))
    rows = list(row_gen)
    assert headers == ["Name", "Vintage"]
    assert len(rows) == 0


# =============================================================================
# XLSX Parsing Tests
# =============================================================================


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Helper to create XLSX bytes from headers and rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_basic() -> None:
    """Test basic XLSX parsing."""
    content = _make_xlsx(
        ["Wine", "Year", "Country"],
        [
            ["Barolo Riserva", 2017, "Italy"],
            ["Rioja Gran Reserva", 2014, "Spain"],
        ],
    )
    headers, row_gen = parse_xlsx(content)
    rows = list(row_gen)
    assert headers == ["Wine", "Year", "Country"]
    assert len(rows) == 2
    assert rows[0]["Wine"] == "Barolo Riserva"
    assert rows[1]["Country"] == "Spain"


def test_parse_xlsx_first_sheet_only() -> None:
    """Test that only the first sheet is parsed."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Name"])
    ws1.append(["First Sheet Wine"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Name"])
    ws2.append(["Second Sheet Wine"])
    buf = io.BytesIO()
    wb.save(buf)

    headers, row_gen = parse_xlsx(buf.getvalue())
    rows = list(row_gen)
    assert len(rows) == 1
    assert rows[0]["Name"] == "First Sheet Wine"


# =============================================================================
# Column Mapping Tests
# =============================================================================


def test_suggest_mapping_exact() -> None:
    """Test exact header matches."""
    mapping = suggest_column_mapping(["name", "winery", "vintage", "country"])
    assert mapping["name"] == "name"
    assert mapping["winery"] == "winery"
    assert mapping["vintage"] == "vintage"
    assert mapping["country"] == "country"


def test_suggest_mapping_aliases() -> None:
    """Test alias matching."""
    mapping = suggest_column_mapping(["Wine Name", "Producer", "Year", "Grape", "Origin"])
    assert mapping["Wine Name"] == "name"
    assert mapping["Producer"] == "winery"
    assert mapping["Year"] == "vintage"
    assert mapping["Grape"] == "grape_variety"
    assert mapping["Origin"] == "country"


def test_suggest_mapping_case_insensitive() -> None:
    """Test case-insensitive matching."""
    mapping = suggest_column_mapping(["WINE NAME", "WINERY", "VINTAGE"])
    assert mapping["WINE NAME"] == "name"
    assert mapping["WINERY"] == "winery"
    assert mapping["VINTAGE"] == "vintage"


def test_suggest_mapping_unknown_custom() -> None:
    """Test that unknown headers default to custom fields using the column name."""
    mapping = suggest_column_mapping(["Cellar Location", "Tasting Room", "name"])
    assert mapping["Cellar Location"] == "custom:Cellar Location"
    assert mapping["Tasting Room"] == "custom:Tasting Room"
    assert mapping["name"] == "name"


# =============================================================================
# Non-Wine Filtering Tests
# =============================================================================


def test_non_wine_whiskey() -> None:
    """Test whiskey row is flagged as non-wine."""
    row = {"Type": "Whiskey", "Name": "Jameson"}
    mapping = {"Type": "wine_type", "Name": "name"}
    assert is_non_wine_row(row, mapping) is True


def test_non_wine_bourbon_in_name() -> None:
    """Test bourbon in name column is flagged."""
    row = {"Type": "Spirit", "Name": "Maker's Mark Bourbon"}
    mapping = {"Type": "wine_type", "Name": "name"}
    assert is_non_wine_row(row, mapping) is True


def test_non_wine_passes_red() -> None:
    """Test that red wine is not flagged."""
    row = {"Type": "Red", "Name": "Chateau Margaux"}
    mapping = {"Type": "wine_type", "Name": "name"}
    assert is_non_wine_row(row, mapping) is False


def test_non_wine_no_type_column() -> None:
    """Test no false positive when there's no type column."""
    row = {"Region": "Bordeaux", "Vintage": "2015"}
    mapping = {"Region": "region", "Vintage": "vintage"}
    assert is_non_wine_row(row, mapping) is False


def test_non_wine_keyword_embedded_in_name() -> None:
    """Test no false positive when keyword is embedded inside a wine name.

    Regression test: 'ale' was matching inside 'NeroBufaleffj' (Gulfi wine).
    """
    row = {"Name": "NeroBufaleffj, Gulfi, Sicily, Italy", "Type": "Red"}
    mapping = {"Name": "name", "Type": "wine_type"}
    assert is_non_wine_row(row, mapping) is False


# =============================================================================
# Row-to-Wine Data Tests
# =============================================================================


def test_row_to_wine_data_basic() -> None:
    """Test basic row conversion."""
    from bson import ObjectId

    row = {"Wine": "Margaux 2015", "Producer": "Chateau Margaux", "Country": "France"}
    mapping = {"Wine": "name", "Producer": "winery", "Country": "country"}
    owner_id = ObjectId()

    result = row_to_wine_data(row, mapping, owner_id)
    assert result is not None
    assert result["name"] == "Margaux 2015"
    assert result["winery"] == "Chateau Margaux"
    assert result["country"] == "France"
    assert result["owner_id"] == owner_id
    assert result["inventory"].quantity == 1


def test_row_to_wine_data_custom_fields() -> None:
    """Test custom fields extraction."""
    from bson import ObjectId

    row = {"Name": "Test Wine", "Location": "Rack 3", "Rating": "95"}
    mapping = {"Name": "name", "Location": "custom:Cellar Location", "Rating": "custom:My Rating"}
    owner_id = ObjectId()

    result = row_to_wine_data(row, mapping, owner_id)
    assert result is not None
    assert result["custom_fields"] == {"Cellar Location": "Rack 3", "My Rating": "95"}
    assert result["custom_fields_text"] is not None
    assert "Cellar Location" in result["custom_fields_text"]
    assert "Rack 3" in result["custom_fields_text"]


def test_row_to_wine_data_no_name_returns_none() -> None:
    """Test that a row without a name returns None."""
    from bson import ObjectId

    row = {"Producer": "Some Winery", "Country": "France"}
    mapping = {"Producer": "winery", "Country": "country"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result is None


def test_row_to_wine_vintage_coercion() -> None:
    """Test vintage year coercion from string."""
    from bson import ObjectId

    row = {"Name": "Test", "Year": "2018"}
    mapping = {"Name": "name", "Year": "vintage"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["vintage"] == 2018


def test_row_to_wine_vintage_float_coercion() -> None:
    """Test vintage coercion from float string (Excel format)."""
    from bson import ObjectId

    row = {"Name": "Test", "Year": "2018.0"}
    mapping = {"Name": "name", "Year": "vintage"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["vintage"] == 2018


def test_row_to_wine_alcohol_coercion() -> None:
    """Test alcohol percentage coercion."""
    from bson import ObjectId

    row = {"Name": "Test", "ABV": "13.5%"}
    mapping = {"Name": "name", "ABV": "alcohol_percentage"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["alcohol_percentage"] == 13.5


def test_row_to_wine_quantity_from_row() -> None:
    """Test quantity taken from row when mapped."""
    from bson import ObjectId

    row = {"Name": "Test", "Qty": "6"}
    mapping = {"Name": "name", "Qty": "quantity"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["inventory"].quantity == 6


def test_row_to_wine_num_cases_multiplies_case_size() -> None:
    """Test num_cases × case_size = total bottles."""
    from bson import ObjectId

    row = {"Name": "Test", "Cases": "2", "Size": "12"}
    mapping = {"Name": "name", "Cases": "num_cases", "Size": "case_size"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["inventory"].quantity == 24  # 2 cases × 12 bottles
    assert result["_num_cases"] == 2
    assert result["_case_size"] == 12


def test_row_to_wine_quantity_with_case_size_derives_cases() -> None:
    """Test quantity (total bottles) + case_size derives case count."""
    from bson import ObjectId

    row = {"Name": "Test", "Qty": "12", "Size": "6"}
    mapping = {"Name": "name", "Qty": "quantity", "Size": "case_size"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["inventory"].quantity == 12  # Quantity IS total bottles
    assert result["_num_cases"] == 2  # 12 / 6 = 2 cases
    assert result["_case_size"] == 6


def test_row_to_wine_case_size_without_quantity() -> None:
    """Test case_size with default quantity of 1 — treated as 1 case."""
    from bson import ObjectId

    row = {"Name": "Test", "Bottles per Case": "6"}
    mapping = {"Name": "name", "Bottles per Case": "case_size"}
    result = row_to_wine_data(row, mapping, ObjectId())
    # Mode 2b: quantity (1) < case_size (6) → treat 1 as number of cases
    assert result["inventory"].quantity == 6  # 1 case × 6 bottles
    assert result["_num_cases"] == 1


def test_row_to_wine_purchase_date() -> None:
    """Test purchase_date parsing from row."""
    from bson import ObjectId

    row = {"Name": "Test", "Date Bought": "2024-03-15"}
    mapping = {"Name": "name", "Date Bought": "purchase_date"}
    result = row_to_wine_data(row, mapping, ObjectId())
    assert result["purchase_date"].year == 2024
    assert result["purchase_date"].month == 3
    assert result["purchase_date"].day == 15


def test_coerce_date_formats() -> None:
    """Test various date formats."""
    from winebox.services.import_service import _coerce_date

    assert _coerce_date("2024-03-15").year == 2024
    assert _coerce_date("15/03/2024").day == 15
    assert _coerce_date("Mar 15, 2024").month == 3
    assert _coerce_date("15 March 2024").day == 15
    assert _coerce_date("") is None
    assert _coerce_date("not a date") is None


def test_wine_identity_key() -> None:
    """Test identity key generation for dedup."""
    from winebox.services.import_service import _wine_identity_key

    wine = {"name": "Chateau Margaux", "winery": "Margaux", "vintage": 2015}
    key = _wine_identity_key(wine)
    assert key == ("chateau margaux", "margaux", "2015")

    # Missing fields use empty string
    wine2 = {"name": "Simple Wine"}
    key2 = _wine_identity_key(wine2)
    assert key2 == ("simple wine", "", "")


def test_duplicate_flag_in_row_to_wine_data() -> None:
    """Test that _duplicate flag is set when wine matches existing set."""
    from bson import ObjectId

    existing = {("test wine", "winery a", "2020")}
    row = {"Name": "Test Wine", "Winery": "Winery A", "Year": "2020"}
    mapping = {"Name": "name", "Winery": "winery", "Year": "vintage"}
    result = row_to_wine_data(row, mapping, ObjectId(), existing_wines=existing)
    assert result.get("_duplicate") is True

    # Non-matching wine should not have flag
    row2 = {"Name": "Different Wine", "Winery": "Winery B", "Year": "2021"}
    result2 = row_to_wine_data(row2, mapping, ObjectId(), existing_wines=existing)
    assert "_duplicate" not in result2


# =============================================================================
# Helper Tests
# =============================================================================


def test_coerce_vintage_valid() -> None:
    assert _coerce_vintage("2020") == 2020


def test_coerce_vintage_float() -> None:
    assert _coerce_vintage("2020.0") == 2020


def test_coerce_vintage_invalid() -> None:
    assert _coerce_vintage("not_a_year") is None


def test_coerce_vintage_out_of_range() -> None:
    assert _coerce_vintage("1800") is None


def test_coerce_float_with_percent() -> None:
    assert _coerce_float("13.5%") == 13.5


def test_coerce_float_empty() -> None:
    assert _coerce_float("") is None


def test_compute_custom_fields_text() -> None:
    result = _compute_custom_fields_text({"Location": "Rack 3", "Price": "$50"})
    assert "Location Rack 3" in result
    assert "Price $50" in result


def test_compute_custom_fields_text_none() -> None:
    assert _compute_custom_fields_text(None) is None
    assert _compute_custom_fields_text({}) is None


# =============================================================================
# AI Column Mapping Tests
# =============================================================================


def _mock_claude_response(content: str) -> MagicMock:
    """Build a mock Anthropic messages.create() response."""
    message = MagicMock()
    message.content = [MagicMock(text=content)]
    return message


@pytest.mark.asyncio
async def test_suggest_mapping_ai_basic() -> None:
    """Test correct mapping returned from mocked Claude response."""
    headers = ["Nom du vin", "Producteur", "Millésime"]
    preview_rows = [
        {"Nom du vin": "Margaux", "Producteur": "Chateau Margaux", "Millésime": "2015"},
    ]
    ai_response = json.dumps({
        "Nom du vin": "name",
        "Producteur": "winery",
        "Millésime": "vintage",
    })

    mock_client = MagicMock()
    mock_client.messages.create = AsyncMock(return_value=_mock_claude_response(ai_response))

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch("anthropic.AsyncAnthropic", return_value=mock_client):
        mock_settings.anthropic_api_key = "test-key"
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is not None
    assert result["Nom du vin"] == "name"
    assert result["Producteur"] == "winery"
    assert result["Millésime"] == "vintage"


@pytest.mark.asyncio
async def test_suggest_mapping_ai_no_api_key() -> None:
    """Test returns None when no API key is set."""
    headers = ["Name", "Vintage"]
    preview_rows = [{"Name": "Test Wine", "Vintage": "2020"}]

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch.dict("os.environ", {}, clear=True):
        mock_settings.anthropic_api_key = None
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is None


@pytest.mark.asyncio
async def test_suggest_mapping_ai_api_error() -> None:
    """Test returns None on API exception (graceful fallback)."""
    headers = ["Name", "Vintage"]
    preview_rows = [{"Name": "Test Wine", "Vintage": "2020"}]

    mock_client = MagicMock()
    mock_client.messages.create = AsyncMock(side_effect=Exception("API rate limit exceeded"))

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch("anthropic.AsyncAnthropic", return_value=mock_client):
        mock_settings.anthropic_api_key = "test-key"
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is None


@pytest.mark.asyncio
async def test_suggest_mapping_ai_malformed_json() -> None:
    """Test returns None when response isn't valid JSON."""
    headers = ["Name", "Vintage"]
    preview_rows = [{"Name": "Test Wine", "Vintage": "2020"}]

    mock_client = MagicMock()
    mock_client.messages.create = AsyncMock(
        return_value=_mock_claude_response("I'm sorry, I can't parse that spreadsheet.")
    )

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch("anthropic.AsyncAnthropic", return_value=mock_client):
        mock_settings.anthropic_api_key = "test-key"
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is None


@pytest.mark.asyncio
async def test_suggest_mapping_ai_invalid_field_dropped() -> None:
    """Test that invalid field values fall back to static per-header."""
    headers = ["Wine Name", "Bogus Column"]
    preview_rows = [{"Wine Name": "Margaux", "Bogus Column": "xyz"}]
    ai_response = json.dumps({
        "Wine Name": "name",
        "Bogus Column": "nonexistent_field",  # Invalid field
    })

    mock_client = MagicMock()
    mock_client.messages.create = AsyncMock(return_value=_mock_claude_response(ai_response))

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch("anthropic.AsyncAnthropic", return_value=mock_client):
        mock_settings.anthropic_api_key = "test-key"
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is not None
    assert result["Wine Name"] == "name"  # Valid AI mapping kept
    # Invalid field falls back to static; "Bogus Column" is not in HEADER_ALIASES
    assert result["Bogus Column"] == "custom:Bogus Column"


@pytest.mark.asyncio
async def test_suggest_mapping_ai_markdown_code_block() -> None:
    """Test handling of JSON wrapped in markdown code blocks."""
    headers = ["Vintge", "Région"]
    preview_rows = [{"Vintge": "2018", "Région": "Bordeaux"}]
    ai_json = json.dumps({"Vintge": "vintage", "Région": "region"})
    wrapped = f"```json\n{ai_json}\n```"

    mock_client = MagicMock()
    mock_client.messages.create = AsyncMock(return_value=_mock_claude_response(wrapped))

    with patch("winebox.services.import_service.mapping.settings") as mock_settings, \
         patch("anthropic.AsyncAnthropic", return_value=mock_client):
        mock_settings.anthropic_api_key = "test-key"
        result = await suggest_column_mapping_ai(headers, preview_rows)

    assert result is not None
    assert result["Vintge"] == "vintage"
    assert result["Région"] == "region"


# =============================================================================
# Static Fallback & Validation Helper Tests
# =============================================================================


def test_static_fallback_known_alias() -> None:
    """Test _static_fallback returns correct field for known alias."""
    assert _static_fallback("Wine Name") == "name"
    assert _static_fallback("Producer") == "winery"
    assert _static_fallback("VINTAGE") == "vintage"


def test_static_fallback_unknown() -> None:
    """Test _static_fallback returns custom field for unknown header."""
    assert _static_fallback("My Custom Column") == "custom:My Custom Column"


def test_is_valid_mapping_value() -> None:
    """Test _is_valid_mapping_value for various inputs."""
    assert _is_valid_mapping_value("name") is True
    assert _is_valid_mapping_value("vintage") is True
    assert _is_valid_mapping_value("skip") is True
    assert _is_valid_mapping_value("custom:Rating") is True
    assert _is_valid_mapping_value("custom:") is False  # Empty custom name
    assert _is_valid_mapping_value("nonexistent_field") is False
    assert _is_valid_mapping_value("") is False


# =============================================================================
# Canonical Fields Tests
# =============================================================================


def test_canonical_fields_subset_of_valid() -> None:
    """Test that every canonical field is a valid wine field."""
    assert set(CANONICAL_WINE_FIELDS).issubset(VALID_WINE_FIELDS)


def test_canonical_fields_name_is_first() -> None:
    """Test that 'name' is the first canonical field (it's required)."""
    assert CANONICAL_WINE_FIELDS[0] == "name"


# =============================================================================
# Generator Contract Tests
# =============================================================================


def test_parse_csv_returns_generator() -> None:
    """Test that parse_csv returns a generator, not a list."""
    from collections.abc import Iterator

    content = "Name,Vintage\nWine A,2020\n"
    headers, row_gen = parse_csv(content.encode("utf-8"))
    assert isinstance(row_gen, Iterator)
    assert not isinstance(row_gen, list)
    # Consuming should still work
    rows = list(row_gen)
    assert len(rows) == 1


def test_parse_xlsx_returns_generator() -> None:
    """Test that parse_xlsx returns a generator, not a list."""
    from collections.abc import Iterator

    content = _make_xlsx(["Name"], [["Wine A"]])
    headers, row_gen = parse_xlsx(content)
    assert isinstance(row_gen, Iterator)
    assert not isinstance(row_gen, list)
    rows = list(row_gen)
    assert len(rows) == 1
