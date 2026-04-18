"""Round-trip tests: import CSV -> export CSV -> verify data matches."""

import csv
import io
from pathlib import Path

import pytest
from httpx import AsyncClient

DATA_DIR = Path(__file__).parent / "data"


def _read_csv_file(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Read a CSV file and return (headers, rows_as_dicts)."""
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = list(reader)
    return list(headers), rows


@pytest.mark.asyncio
async def test_custom_fields_roundtrip(client: AsyncClient) -> None:
    """Import a CSV with custom fields, export CSV, verify custom fields are expanded."""
    csv_path = DATA_DIR / "roundtrip-test-data.csv"
    original_headers, original_rows = _read_csv_file(csv_path)

    # 1. Upload CSV
    with open(csv_path, "rb") as f:
        files = {"file": ("roundtrip-test-data.csv", f, "text/csv")}
        upload_resp = await client.post("/api/import/upload", files=files)
    assert upload_resp.status_code == 200
    batch_id = upload_resp.json()["batch_id"]

    # 2. Set column mapping: map known fields, rest to custom:*
    mapping = {
        "Description": "name",
        "Country": "country",
        "Region": "region",
        "Vintage": "vintage",
        "Colour": "wine_type",
        "Quantity in Bottles": "quantity",
        # Custom fields
        "Parent ID": "custom:Parent ID",
        "Product Code(s)": "custom:Product Code(s)",
        "Maturity": "custom:Maturity",
        "Bottle Format": "custom:Bottle Format",
        "Bottle Volume": "custom:Bottle Volume",
        "Current Status": "custom:Current Status",
        "Provenance": "custom:Provenance",
    }
    map_resp = await client.post(
        f"/api/import/{batch_id}/mapping",
        json={"mapping": mapping},
    )
    assert map_resp.status_code == 200

    # 3. Process import
    process_resp = await client.post(
        f"/api/import/{batch_id}/process",
        json={"skip_non_wine": False, "default_quantity": 1},
    )
    assert process_resp.status_code == 200
    result = process_resp.json()
    assert result["wines_created"] == len(original_rows)
    assert result["status"] == "completed"

    # 4. Export as CSV
    export_resp = await client.get("/api/export/wines?format=csv")
    assert export_resp.status_code == 200

    exported_content = export_resp.content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(exported_content))
    exported_headers = reader.fieldnames or []
    exported_rows = list(reader)

    # 5. Verify exported row count matches
    assert len(exported_rows) == len(original_rows)

    # 6. Verify custom field columns exist as individual headers
    custom_field_names = [
        "Bottle Format", "Bottle Volume", "Current Status",
        "Maturity", "Parent ID", "Product Code(s)", "Provenance",
    ]
    for cf in custom_field_names:
        assert cf in exported_headers, f"Custom field '{cf}' not in exported headers"

    # 7. Verify 'custom_fields' is NOT a column (no longer a single JSON blob)
    assert "custom_fields" not in exported_headers

    # 8. Build a lookup of exported wines by name for comparison
    exported_by_name: dict[str, dict[str, str]] = {}
    for row in exported_rows:
        exported_by_name[row["name"]] = row

    # 9. Verify each imported row round-trips correctly
    for orig in original_rows:
        wine_name = orig["Description"]
        assert wine_name in exported_by_name, f"Wine '{wine_name}' missing from export"
        exported = exported_by_name[wine_name]

        # Standard fields
        assert exported["country"] == orig["Country"]
        assert exported["region"] == orig["Region"]
        assert exported["vintage"] == orig["Vintage"]
        assert exported["wine_type"] == orig["Colour"]
        assert exported["quantity"] == orig["Quantity in Bottles"]

        # Custom fields — each should be its own column
        assert exported["Parent ID"] == orig["Parent ID"]
        assert exported["Product Code(s)"] == orig["Product Code(s)"]
        assert exported["Maturity"] == orig["Maturity"]
        assert exported["Bottle Format"] == orig["Bottle Format"]
        assert exported["Bottle Volume"] == orig["Bottle Volume"]
        assert exported["Current Status"] == orig["Current Status"]
        assert exported["Provenance"] == orig["Provenance"]


@pytest.mark.asyncio
async def test_export_csv_no_custom_fields_still_works(
    client: AsyncClient, sample_image_bytes: bytes,
) -> None:
    """Export works correctly when wines have no custom fields."""
    # Create a wine without custom fields
    files = {
        "front_label": ("test.png", io.BytesIO(sample_image_bytes), "image/png"),
    }
    data = {"name": "Plain Wine", "country": "Spain", "quantity": "2"}
    await client.post("/api/wines/checkin", files=files, data=data)

    export_resp = await client.get("/api/export/wines?format=csv")
    assert export_resp.status_code == 200

    content = export_resp.content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = list(reader)

    # No custom fields columns should appear
    assert "custom_fields" not in headers
    assert len(rows) == 1
    assert rows[0]["name"] == "Plain Wine"
    assert rows[0]["country"] == "Spain"


@pytest.mark.asyncio
async def test_export_xlsx_custom_fields_expanded(client: AsyncClient) -> None:
    """XLSX export also expands custom fields into individual columns."""
    from openpyxl import load_workbook

    # Import CSV with custom fields
    csv_content = (
        "Name,Location,Purchase Price\n"
        "Test Cab Sauv,Rack 1A,$45\n"
        "Test Merlot,Rack 2B,$30\n"
    )
    files = {"file": ("test.csv", io.BytesIO(csv_content.encode()), "text/csv")}
    upload_resp = await client.post("/api/import/upload", files=files)
    assert upload_resp.status_code == 200
    batch_id = upload_resp.json()["batch_id"]

    mapping = {
        "Name": "name",
        "Location": "custom:Location",
        "Purchase Price": "custom:Purchase Price",
    }
    await client.post(
        f"/api/import/{batch_id}/mapping",
        json={"mapping": mapping},
    )
    await client.post(
        f"/api/import/{batch_id}/process",
        json={"skip_non_wine": False, "default_quantity": 1},
    )

    # Export as XLSX
    export_resp = await client.get("/api/export/wines?format=xlsx")
    assert export_resp.status_code == 200

    wb = load_workbook(io.BytesIO(export_resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    assert "Location" in headers
    assert "Purchase Price" in headers
    assert "custom_fields" not in headers

    # Check data values
    loc_col = headers.index("Location") + 1
    price_col = headers.index("Purchase Price") + 1

    row_values: dict[str, dict[str, str]] = {}
    name_col = headers.index("name") + 1
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        row_values[name] = {
            "Location": str(ws.cell(row=row_idx, column=loc_col).value or ""),
            "Purchase Price": str(ws.cell(row=row_idx, column=price_col).value or ""),
        }

    assert row_values["Test Cab Sauv"]["Location"] == "Rack 1A"
    assert row_values["Test Cab Sauv"]["Purchase Price"] == "$45"
    assert row_values["Test Merlot"]["Location"] == "Rack 2B"
    assert row_values["Test Merlot"]["Purchase Price"] == "$30"


@pytest.mark.asyncio
async def test_mixed_custom_fields_across_wines(client: AsyncClient) -> None:
    """Wines with different custom fields get all columns, empty where absent."""
    # First import: wine with Location
    csv1 = "Name,Location\nWine Alpha,Cellar A\n"
    files1 = {"file": ("w1.csv", io.BytesIO(csv1.encode()), "text/csv")}
    resp1 = await client.post("/api/import/upload", files=files1)
    batch1 = resp1.json()["batch_id"]
    await client.post(
        f"/api/import/{batch1}/mapping",
        json={"mapping": {"Name": "name", "Location": "custom:Location"}},
    )
    await client.post(
        f"/api/import/{batch1}/process",
        json={"skip_non_wine": False, "default_quantity": 1},
    )

    # Second import: wine with Rack Number
    csv2 = "Name,Rack Number\nWine Beta,R3\n"
    files2 = {"file": ("w2.csv", io.BytesIO(csv2.encode()), "text/csv")}
    resp2 = await client.post("/api/import/upload", files=files2)
    batch2 = resp2.json()["batch_id"]
    await client.post(
        f"/api/import/{batch2}/mapping",
        json={"mapping": {"Name": "name", "Rack Number": "custom:Rack Number"}},
    )
    await client.post(
        f"/api/import/{batch2}/process",
        json={"skip_non_wine": False, "default_quantity": 1},
    )

    # Export CSV
    export_resp = await client.get("/api/export/wines?format=csv")
    content = export_resp.content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = list(reader)

    # Both custom field columns should exist
    assert "Location" in headers
    assert "Rack Number" in headers

    by_name = {r["name"]: r for r in rows}

    # Wine Alpha has Location but not Rack Number
    assert by_name["Wine Alpha"]["Location"] == "Cellar A"
    assert by_name["Wine Alpha"]["Rack Number"] == ""

    # Wine Beta has Rack Number but not Location
    assert by_name["Wine Beta"]["Rack Number"] == "R3"
    assert by_name["Wine Beta"]["Location"] == ""
